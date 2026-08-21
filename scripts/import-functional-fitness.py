#!/usr/bin/env python3
"""
Import Functional Fitness Exercise Database → data/exercises-functional.json.

Reads the .xlsx export of the source Google Sheet. XLSX is required (not CSV)
because Google Sheets strips hyperlink URLs on CSV export — the YouTube URLs
live in `=HYPERLINK(...)` cells and only survive the XLSX export.

Usage:
    python3 scripts/import-functional-fitness.py [path-to-xlsx]

Defaults to data/functional-fitness.xlsx. To refresh from the sheet:
    curl -sSL 'https://docs.google.com/spreadsheets/d/<SHEET_ID>/export?format=xlsx' \\
      -o data/functional-fitness.xlsx
    python3 scripts/import-functional-fitness.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

# openpyxl uses 1-indexed columns. Row 16 is the header; data starts at row 17.
# The sheet has an empty column A (index 1) before the real data.
COL = {
    'name':                 2,
    'short_yt':             3,
    'deep_yt':              4,
    'difficulty':           5,
    'target':               6,
    'prime_mover':          7,
    'secondary_muscle':     8,
    'tertiary_muscle':      9,
    'primary_equipment':    10,
    'secondary_equipment':  12,
    'posture':              14,
    'grip':                 17,
    'movement_pattern_1':   22,
    'movement_pattern_2':   23,
    'movement_pattern_3':   24,
    'plane_of_motion_1':    25,
    'plane_of_motion_2':    26,
    'plane_of_motion_3':    27,
    'body_region':          28,
    'force_type':           29,
    'mechanics':            30,
    'laterality':           31,
    'classification':       32,
}

BODY_REGION_TO_PART = {
    'Core':        'waist',
    'Lower Body':  'upper legs',
    'Upper Body':  'chest',
    'Full Body':   'cardio',
}


def norm(s) -> str:
    return re.sub(r'\s+', ' ', (str(s) if s is not None else '')).strip()


def clean_choice(s) -> str:
    return norm(s).rstrip('*').strip()


def cell(ws, row: int, key: str):
    return ws.cell(row=row, column=COL[key])


def transform(ws, row_num: int, out_idx: int) -> dict:
    def val(key: str) -> str:
        return norm(cell(ws, row_num, key).value)

    def hlink(key: str) -> str:
        c = cell(ws, row_num, key)
        return c.hyperlink.target if c.hyperlink else ''

    name = val('name')
    body_region = clean_choice(val('body_region'))
    target = clean_choice(val('target')).lower()
    prime = clean_choice(val('prime_mover')).lower()
    secondary = [x for x in (clean_choice(val('secondary_muscle')).lower(),
                             clean_choice(val('tertiary_muscle')).lower()) if x]
    equipment = clean_choice(val('primary_equipment')).lower()
    movement_patterns = [p for p in (val('movement_pattern_1'),
                                     val('movement_pattern_2'),
                                     val('movement_pattern_3')) if p]
    planes = [p for p in (val('plane_of_motion_1'),
                          val('plane_of_motion_2'),
                          val('plane_of_motion_3')) if p]
    body_part = BODY_REGION_TO_PART.get(body_region, '')

    return {
        'id': f'ff-{out_idx:04d}',
        'name': name.lower(),
        'category': body_part,
        'body_part': body_part,
        'body_region': body_region,
        'equipment': equipment,
        'target': target,
        'muscle_group': prime,
        'secondary_muscles': secondary,
        'instructions': '',
        'instruction_steps': [],
        'media_id': '',
        'image': '',
        'gif_url': '',
        'video_url':       hlink('short_yt'),
        'explanation_url': hlink('deep_yt'),
        'attribution': '',
        'has_media': False,
        'classification':      clean_choice(val('classification')),
        'difficulty':          clean_choice(val('difficulty')),
        'mechanics':           clean_choice(val('mechanics')),
        'laterality':          clean_choice(val('laterality')),
        'force_type':          clean_choice(val('force_type')),
        'posture':             clean_choice(val('posture')),
        'grip':                clean_choice(val('grip')),
        'movement_patterns':   movement_patterns,
        'planes_of_motion':    planes,
    }


def main() -> None:
    default_path = Path(__file__).resolve().parent.parent / 'data' / 'functional-fitness.xlsx'
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    out_path = Path(__file__).resolve().parent.parent / 'data' / 'exercises-functional.json'

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    ws = wb.active

    records: list[dict] = []
    # Data starts at row 17; iterate through all rows with a name.
    for row_num in range(17, ws.max_row + 1):
        name = norm(cell(ws, row_num, 'name').value)
        if not name:
            continue
        records.append(transform(ws, row_num, len(records) + 1))

    out_path.write_text(json.dumps(records, ensure_ascii=False, indent=None), encoding='utf-8')

    from collections import Counter
    def top(field: str, n: int = 5) -> str:
        c = Counter(r[field] for r in records if r[field])
        return ', '.join(f'{k}={v}' for k, v in c.most_common(n))

    with_video = sum(1 for r in records if r['video_url'])
    with_explanation = sum(1 for r in records if r['explanation_url'])

    print(f'Wrote {len(records)} records → {out_path.relative_to(Path.cwd())}')
    print(f'  YouTube video_url populated:       {with_video} / {len(records)}')
    print(f'  YouTube explanation_url populated: {with_explanation} / {len(records)}')
    print(f'  classifications: {top("classification")}')
    print(f'  difficulties:    {top("difficulty")}')
    print(f'  body_regions:    {top("body_region")}')


if __name__ == '__main__':
    main()
